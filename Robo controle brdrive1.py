import os
import sys
import io
import unicodedata

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import difflib
import subprocess
import pandas as pd
import requests
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURAÇÕES
# ─────────────────────────────────────────────

ARQUIVO_LOCAL = r"C:\Users\robson.noberto\Desktop\Power BI Gus\BRDrive_BI_Novo.xlsx"

SHAREPOINT_URL = "https://spjtexpress-my.sharepoint.com/:x:/g/personal/gerson_silva_spjtexpress_onmicrosoft_com/IQADZHQau4f1T7KrFoI4SqRdAV-wnc4VJ_-s2J5BbqNT5qw?e=p0J8tR"

ONEDRIVE_PASTA = None  # deixe None para detectar automaticamente

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Marco", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
}

# ─────────────────────────────────────────────
# DETECTAR PASTA DO ONEDRIVE
# ─────────────────────────────────────────────

def detectar_onedrive():
    if ONEDRIVE_PASTA and os.path.exists(ONEDRIVE_PASTA):
        return ONEDRIVE_PASTA

    usuario = os.environ.get("USERNAME", "robson.noberto")

    candidatos = [
        rf"C:\Users\{usuario}\OneDrive - J&T EXPRESS - FILIAL SP",
        rf"C:\Users\{usuario}\OneDrive - J&T EXPRESS",
        rf"C:\Users\{usuario}\OneDrive - SPJTEXPRESS",
        rf"C:\Users\{usuario}\OneDrive - spjtexpress",
        rf"C:\Users\{usuario}\OneDrive - JT Express",
        rf"C:\Users\{usuario}\OneDrive",
    ]

    for caminho in candidatos:
        if os.path.exists(caminho):
            return caminho

    for var in ["OneDriveCommercial", "OneDriveConsumer", "OneDrive"]:
        val = os.environ.get(var)
        if val and os.path.exists(val):
            return val

    return None


# ─────────────────────────────────────────────
# PASSO 1 — Calcular KPIs de TODOS os meses
# ─────────────────────────────────────────────

def calcular_kpis_por_mes(arquivo_local):
    print(">> Lendo arquivo BRDrive...")
    df = pd.read_excel(arquivo_local, sheet_name='fPrincipal')

    jet = df[df['Transportador'].str.strip().str.upper() == 'JET SP'].copy()
    jet['DATA'] = pd.to_datetime(jet['DATA'])
    jet['ANO']  = jet['DATA'].dt.year
    jet['MES']  = jet['DATA'].dt.month

    meses_disponiveis = sorted(jet[['ANO', 'MES']].drop_duplicates().itertuples(index=False))
    meses_disponiveis = [m for m in meses_disponiveis if m.ANO >= 2026]
    print(f">> Meses encontrados: {[f'{MESES_PT[m.MES]}/{m.ANO}' for m in meses_disponiveis]}")

    resultado_meses = {}

    for periodo_info in meses_disponiveis:
        ano, mes = periodo_info.ANO, periodo_info.MES
        mes_nome = MESES_PT[mes]
        grupo_mes = jet[(jet['ANO'] == ano) & (jet['MES'] == mes)].copy()

        resultado = []
        for condutor, grp in grupo_mes.groupby('CONDUTOR'):
            tot = len(grp)
            if tot == 0:
                continue

            # UTILIZACAO
            on_s = (grp['Tempo Saida OFF'].fillna('OFF').str.strip().str.upper() != 'OFF').sum()
            on_c = (grp['Tempo chegada OFF'].fillna('OFF').str.strip().str.upper() != 'OFF').sum()
            kpi_util_s = round(on_s / tot * 100, 1)
            kpi_util_c = round(on_c / tot * 100, 1)
            utilizacao = round((kpi_util_s + kpi_util_c) / 2, 1)

            # PONTUALIDADE
            com_s = grp['PONTUALIDADE SAÍDA'].notna().sum()
            com_c = grp['PONTUALIDADE CHEGADA'].notna().sum()
            np_s = (grp['PONTUALIDADE SAÍDA'].str.strip().str.lower() == 'no prazo').sum()
            np_c = (grp['PONTUALIDADE CHEGADA'].str.strip().str.lower() == 'no prazo').sum()
            kpi_pont_s = round(np_s / com_s * 100, 1) if com_s > 0 else 0
            kpi_pont_c = round(np_c / com_c * 100, 1) if com_c > 0 else 0
            pontualidade = round((kpi_pont_s + kpi_pont_c) / 2, 1)

            resultado.append({
                'CONDUTOR': condutor.strip().upper(),
                'UTILIZACAO': utilizacao,
                'PONTUALIDADE': pontualidade
            })

        kpis = pd.DataFrame(resultado)
        print(f"  [OK] {mes_nome}/{ano}: KPIs de {len(kpis)} motoristas")
        resultado_meses[(ano, mes)] = (kpis, mes_nome, ano)

    return resultado_meses


# ─────────────────────────────────────────────
# PASSO 2 — Baixar arquivo do SharePoint
# ─────────────────────────────────────────────

def baixar_sharepoint(url, destino):
    print(">> Baixando Controle BRDRIVE do SharePoint...")

    download_url = url + ("&download=1" if "?" in url else "?download=1")
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    try:
        resp = requests.get(download_url, headers=headers, allow_redirects=True, timeout=30)
        content_type = resp.headers.get('Content-Type', '')
        is_excel = any(x in content_type for x in ['spreadsheet', 'octet-stream', 'excel'])
        if resp.status_code == 200 and len(resp.content) > 5000 and is_excel:
            with open(destino, 'wb') as f:
                f.write(resp.content)
            print(f"[OK] Arquivo baixado ({len(resp.content)/1024:.0f} KB)")
            return True
    except Exception as e:
        print(f"[AVISO] Erro ao baixar: {e}")

    print("[AVISO] Nao foi possivel baixar automaticamente.")
    print(f"   Baixe o arquivo manualmente e salve como:\n   {destino}")
    return False


# ─────────────────────────────────────────────
# UTILITÁRIOS DE NOME
# ─────────────────────────────────────────────

def normalizar(nome):
    sem_acento = unicodedata.normalize('NFD', nome)
    sem_acento = ''.join(c for c in sem_acento if unicodedata.category(c) != 'Mn')
    return ' '.join(sem_acento.upper().split())

def buscar_kpi(nome, kpi_dict, kpi_dict_norm, limite=0.82):
    chave = nome.upper()
    if chave in kpi_dict:
        return kpi_dict[chave], None

    chave_norm = normalizar(nome)
    if chave_norm in kpi_dict_norm:
        match = kpi_dict_norm[chave_norm]
        return kpi_dict[match], match

    proximos = difflib.get_close_matches(chave_norm, list(kpi_dict_norm.keys()), n=1, cutoff=limite)
    if proximos:
        match = kpi_dict_norm[proximos[0]]
        return kpi_dict[match], match

    palavras = set(chave_norm.split())
    melhor_match, melhor_score = None, 0
    for candidato in kpi_dict_norm:
        palavras_cand = set(candidato.split())
        if palavras.issubset(palavras_cand):
            score = len(palavras) / len(palavras_cand)
            if score > melhor_score:
                melhor_score = score
                melhor_match = candidato
    if melhor_match:
        match = kpi_dict_norm[melhor_match]
        return kpi_dict[match], match

    return None, None

def cor_kpi(valor_decimal, meta=0.95):
    if valor_decimal is None:
        return "000000"
    if valor_decimal >= meta:
        return "27AE60"
    elif valor_decimal >= meta - 0.15:
        return "E67E22"
    else:
        return "E74C3C"


# ─────────────────────────────────────────────
# MÊS ESPELHO — lógica de troca
# ─────────────────────────────────────────────

def determinar_mes_espelho(hoje):
    """
    Dia > 15 → usa o mês atual.
    Dia <= 15 → usa o mês anterior (dados ainda incompletos).
    """
    if hoje.day > 15:
        return hoje.month, hoje.year
    else:
        if hoje.month == 1:
            return 12, hoje.year - 1
        return hoje.month - 1, hoje.year


def espelhar_planilha1(arquivo_excel, kpis, mes_nome, ano):
    """Atualiza colunas D e E da Planilha1 com os KPIs do mês espelho.
    - Atualiza D e E dos motoristas já presentes (sem mover linhas)
    - Adiciona ao FINAL os que estão na base (司机打卡控制表) mas faltam na Planilha1
    - Nunca remove linhas nem altera colunas além de D e E"""
    wb = load_workbook(arquivo_excel)
    if 'Planilha1' not in wb.sheetnames:
        print("[AVISO] Planilha1 nao encontrada — espelho nao aplicado.")
        return

    ws_p1 = wb['Planilha1']
    kpi_dict = {row['CONDUTOR']: row for _, row in kpis.iterrows()}
    kpi_dict_norm = {normalizar(k): k for k in kpi_dict}

    font_verde    = Font(color="27AE60", bold=True)
    font_vermelho = Font(color="E74C3C", bold=True)
    font_normal   = Font(color="000000")

    # Passo 1 — nomes já presentes na Planilha1 (normalizados, para comparação)
    nomes_existentes = set()
    for row in ws_p1.iter_rows(values_only=True):
        if row[0] and not any(p in str(row[0]).upper() for p in ["NOME", "姓名", "NAME"]):
            nomes_existentes.add(normalizar(str(row[0]).strip()))

    # Passo 2 — atualiza D e E para quem já está na Planilha1
    atualizados, sem_dados = 0, []
    for row in ws_p1.iter_rows(min_row=1):
        if not row[0].value:
            continue
        nome = str(row[0].value).strip()
        if any(p in nome.upper() for p in ["NOME", "姓名", "NAME"]):
            continue
        kpi_row, _ = buscar_kpi(nome, kpi_dict, kpi_dict_norm)
        row[3].number_format = '0.00%'
        row[4].number_format = '0.00%'
        if kpi_row is not None:
            val_u = round(kpi_row['UTILIZACAO'] / 100, 4)
            val_p = round(kpi_row['PONTUALIDADE'] / 100, 4)
            row[3].value = val_u
            row[4].value = val_p
            row[3].font = font_verde if val_u >= 0.95 else font_vermelho
            row[4].font = font_verde if val_p >= 0.95 else font_vermelho
            atualizados += 1
        else:
            row[3].value = None
            row[4].value = None
            row[3].font = font_normal
            row[4].font = font_normal
            sem_dados.append(nome)

    # Passo 3 — adiciona ao FINAL quem está na base mas falta na Planilha1
    aba_base = next((s for s in wb.sheetnames if '司机打卡控制表' in s), None)
    adicionados = 0
    if aba_base:
        ws_base = wb[aba_base]
        proxima = ws_p1.max_row + 1
        for row_b in ws_base.iter_rows(values_only=True):
            if not row_b[0]:
                continue
            nome = str(row_b[0]).strip()
            if any(p in nome.upper() for p in ["NOME", "姓名", "NAME", "BRDRIVE", "控制表"]):
                continue
            nome_norm = normalizar(nome)
            # Verifica se já existe (exato ou similar)
            if nome_norm in nomes_existentes:
                continue
            if difflib.get_close_matches(nome_norm, nomes_existentes, n=1, cutoff=0.82):
                continue
            # Não encontrado: adiciona linha nova ao final
            rg  = row_b[1] if len(row_b) > 1 else None
            cpf = row_b[2] if len(row_b) > 2 else None
            kpi_row, _ = buscar_kpi(nome, kpi_dict, kpi_dict_norm)
            ws_p1.cell(row=proxima, column=1).value = nome
            ws_p1.cell(row=proxima, column=2).value = rg
            ws_p1.cell(row=proxima, column=3).value = cpf
            cel_d = ws_p1.cell(row=proxima, column=4)
            cel_e = ws_p1.cell(row=proxima, column=5)
            cel_d.number_format = '0.00%'
            cel_e.number_format = '0.00%'
            if kpi_row is not None:
                val_u = round(kpi_row['UTILIZACAO'] / 100, 4)
                val_p = round(kpi_row['PONTUALIDADE'] / 100, 4)
                cel_d.value = val_u
                cel_e.value = val_p
                cel_d.font = font_verde if val_u >= 0.95 else font_vermelho
                cel_e.font = font_verde if val_p >= 0.95 else font_vermelho
            else:
                cel_d.value = None
                cel_e.value = None
                cel_d.font = font_normal
                cel_e.font = font_normal
            nomes_existentes.add(nome_norm)
            proxima += 1
            adicionados += 1

    wb.save(arquivo_excel)
    total = atualizados + adicionados
    print(f"[OK] Planilha1 espelhada com {mes_nome}/{ano}: {total} motoristas ({adicionados} adicionados ao final)")
    if sem_dados:
        nomes_str = ', '.join(sem_dados[:3])
        sufixo = '...' if len(sem_dados) > 3 else ''
        print(f"   Sem dados BRDrive ({len(sem_dados)}): {nomes_str}{sufixo}")


# ─────────────────────────────────────────────
# PASSO 3 — Criar aba para cada mês
# ─────────────────────────────────────────────

def criar_nova_aba(arquivo_excel, kpis, mes_nome, ano):
    nome_aba = f"KPIs {mes_nome} {ano}"
    print(f">> Criando aba '{nome_aba}'...")

    wb = load_workbook(arquivo_excel)

    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    aba_base = 'Folha1' if 'Folha1' in wb.sheetnames else wb.sheetnames[0]
    ws_base = wb[aba_base]
    ws_nova = wb.create_sheet(nome_aba)

    fills_header = {
        'vermelho': PatternFill("solid", fgColor="CC0000"),
        'laranja':  PatternFill("solid", fgColor="C55A11"),
        'rosa':     PatternFill("solid", fgColor="C00066"),
        'azul':     PatternFill("solid", fgColor="0070C0"),
    }
    fill_par   = PatternFill("solid", fgColor="F2F2F2")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")
    font_cab   = Font(name='Arial', bold=True, color="FFFFFF", size=10)
    font_nome  = Font(name='Arial', size=9)
    font_valor = Font(name='Arial', bold=True, size=9)
    font_kpi_verde    = Font(name='Arial', bold=True, size=9, color="27AE60")
    font_kpi_vermelho = Font(name='Arial', bold=True, size=9, color="E74C3C")
    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    esq    = Alignment(horizontal='left',   vertical='center')
    borda  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    colunas = [
        ("Nome  姓名",              'vermelho', 42),
        ("No RG   RG 号",           'vermelho', 16),
        ("CPF 公积金",               'vermelho', 18),
        ("Utilizacao % 利用率%",     'laranja',  20),
        ("Pontualidade % 准时率%",   'rosa',     22),
        ("Assiduidade % 出勤率",     'azul',     22),
    ]

    for col_idx, (titulo, fill_key, largura) in enumerate(colunas, start=1):
        cell = ws_nova.cell(row=1, column=col_idx, value=titulo)
        cell.fill = fills_header[fill_key]
        cell.font = font_cab
        cell.alignment = centro
        cell.border = borda
        ws_nova.column_dimensions[get_column_letter(col_idx)].width = largura
    ws_nova.row_dimensions[1].height = 32

    kpi_dict = {row['CONDUTOR']: row for _, row in kpis.iterrows()}
    kpi_dict_norm = {normalizar(k): k for k in kpi_dict}

    dados_base = []
    for row in ws_base.iter_rows(values_only=True):
        if not row[0]:
            continue

        nome = str(row[0]).strip().upper()

        # ignora cabeçalho e título (PT + CH + variações)
        if any(p in nome for p in ["NOME", "姓名", "NAME", "BRDRIVE", "控制表"]):
            continue

        dados_base.append((str(row[0]).strip(), row[1], row[2]))

    sem_dados = []
    for i, (nome, rg, cpf) in enumerate(dados_base, start=2):
        fill_linha = fill_par if i % 2 == 0 else fill_impar
        kpi_row, nome_match = buscar_kpi(nome, kpi_dict, kpi_dict_norm)

        if nome_match:
            print(f"   -> '{nome}' => '{nome_match}' (nome similar)")
        if kpi_row is None:
            sem_dados.append(nome)

        util = kpi_row['UTILIZACAO']   / 100 if kpi_row is not None else None
        pont = kpi_row['PONTUALIDADE'] / 100 if kpi_row is not None else None

        for col_idx, valor in enumerate([nome, rg, cpf, util, pont, None], start=1):
            cell = ws_nova.cell(row=i, column=col_idx, value=valor)
            cell.border = borda
            cell.fill   = fill_linha
            if col_idx == 1:
                cell.font = font_nome
                cell.alignment = esq
            elif col_idx in (4, 5):
                cell.number_format = '0.00%'
                cell.alignment = centro
                if valor is not None:
                    cell.font = font_kpi_verde if valor >= 0.95 else font_kpi_vermelho
                else:
                    cell.font = font_valor
            elif col_idx == 6:
                cell.number_format = '0.00%'
                cell.alignment = centro
                cell.font = font_valor
            else:
                cell.font = font_nome
                cell.alignment = centro
        ws_nova.row_dimensions[i].height = 18

    ws_nova.freeze_panes = 'A2'
    wb.save(arquivo_excel)

    encontrados = len(dados_base) - len(sem_dados)
    print(f"[OK] Aba '{nome_aba}' criada! ({encontrados}/{len(dados_base)} motoristas com KPI)")
    if sem_dados:
        print(f"   {len(sem_dados)} motorista(s) sem dados no BRDrive:")
        for n in sem_dados:
            print(f"      [X] {n}")
    return nome_aba


# ─────────────────────────────────────────────
# PASSO 4 — Gerar arquivo RH (mes atual)
# ─────────────────────────────────────────────

def gerar_arquivo_rh(arquivo_base, kpis, mes_nome, ano):
    import shutil
    import tempfile

    print(f">> Gerando arquivo RH para {mes_nome}/{ano}...")

    # Cria uma copia temporaria do arquivo base (SharePoint)
    tmp = os.path.join(tempfile.gettempdir(), f"_rh_temp_{mes_nome}_{ano}.xlsx")
    shutil.copy2(arquivo_base, tmp)

    # Cria a aba do mes atual nessa copia
    criar_nova_aba(tmp, kpis, mes_nome, ano)

    # Remove abas extras (deixa apenas Folha1 e a aba do mes)
    wb = load_workbook(tmp)
    nome_aba = f"KPIs {mes_nome} {ano}"
    for sheet in wb.sheetnames:
        if sheet != nome_aba and sheet != 'Folha1':
            del wb[sheet]
    wb.save(tmp)

    return tmp


# ─────────────────────────────────────────────
# PASSO 5 — Salvar no OneDrive local
# ─────────────────────────────────────────────

def salvar_onedrive(arquivo_consolidado, arquivo_rh, mes_nome, ano):
    import shutil
    onedrive = detectar_onedrive()

    subpasta = os.path.join(
        onedrive if onedrive else os.path.expanduser("~") + "\\Desktop",
        "Controle KPIs Motoristas"
    )
    os.makedirs(subpasta, exist_ok=True)

    # Consolidado anual
    nome_consolidado = f"Controle BRDRIVE - KPIs {ano}.xlsx"
    destino_consolidado = os.path.join(subpasta, nome_consolidado)
    shutil.copy2(arquivo_consolidado, destino_consolidado)

    # Arquivo RH do mes atual
    nome_rh = f"Controle BRDRIVE - KPIs {mes_nome} {ano} - RH.xlsx"
    destino_rh = os.path.join(subpasta, nome_rh)
    shutil.copy2(arquivo_rh, destino_rh)

    print()
    print("[OK] Arquivos salvos no OneDrive!")
    print(f"   Pasta: {subpasta}")
    print(f"   [1] {nome_consolidado}  (consolidado todos os meses)")
    print(f"   [2] {nome_rh}  (envio ao RH)")

    if not onedrive:
        print()
        print("[AVISO] OneDrive nao encontrado — arquivos salvos no Desktop.")
    else:
        print()
        print("   Os arquivos serao sincronizados automaticamente com a nuvem.")

    subprocess.Popen(f'explorer "{subpasta}"')


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print()
    print("=" * 55)
    print("   ROBO CONTROLE BRDRIVE - JET SP")
    print("=" * 55)
    print()

    if not os.path.exists(ARQUIVO_LOCAL):
        print(f"[ERRO] Arquivo nao encontrado:\n   {ARQUIVO_LOCAL}")
        sys.exit(1)

    hoje = date.today()

    # 🔒 mês anterior (fechado)
    if hoje.month == 1:
        mes_ref = 12
        ano_ref = hoje.year - 1
    else:
        mes_ref = hoje.month - 1
        ano_ref = hoje.year

    mes_nome_ref = MESES_PT[mes_ref]

    # 🔹 arquivo continua sendo do ano atual (consolidado)
    arquivo_temp = os.path.join(
        os.path.expanduser("~"), "Desktop",
        f"Controle_BRDRIVE_KPIs_{hoje.year}.xlsx"
    )

    # Passo 1: KPIs de todos os meses
    todos_kpis = calcular_kpis_por_mes(ARQUIVO_LOCAL)

    # Lê TODAS as abas locais (com fórmulas) ANTES do download
    # Garante que nenhuma fórmula ou dado manual seja perdido ao sobrescrever
    onedrive_pre = detectar_onedrive()
    abas_locais = {}  # { nome_aba: [[cell.value, ...], ...] }
    if onedrive_pre:
        arquivo_local_ctrl = os.path.join(
            onedrive_pre, "Controle KPIs Motoristas",
            f"Controle BRDRIVE - KPIs {hoje.year}.xlsx"
        )
        if os.path.exists(arquivo_local_ctrl):
            try:
                wb_local = load_workbook(arquivo_local_ctrl, data_only=False)
                for nome_aba in wb_local.sheetnames:
                    if not nome_aba.startswith('KPIs '):
                        ws_l = wb_local[nome_aba]
                        abas_locais[nome_aba] = [
                            [cell.value for cell in row]
                            for row in ws_l.iter_rows()
                        ]
                nomes = list(abas_locais.keys())
                print(f">> Abas locais preservadas (com formulas): {nomes}")
            except Exception as e:
                print(f"[AVISO] Nao foi possivel ler abas locais: {e}")

    # Passo 2: Baixar SharePoint
    if not baixar_sharepoint(SHAREPOINT_URL, arquivo_temp):
        sys.exit(1)

    # Passo 3: Determina o mês espelho e mantém só ele no consolidado
    mes_esp, ano_esp = determinar_mes_espelho(hoje)
    mes_nome_esp = MESES_PT[mes_esp]
    print()
    print(f">> Mes espelho: {mes_nome_esp}/{ano_esp}  (regra: dia {hoje.day} {'> 15 → mes atual' if hoje.day > 15 else '<= 15 → mes anterior'})")

    # Remove abas de KPI antigas do arquivo baixado
    wb_tmp = load_workbook(arquivo_temp)
    abas_remover = [s for s in wb_tmp.sheetnames if s.startswith('KPIs ')]
    for aba in abas_remover:
        del wb_tmp[aba]
        print(f"   Removida aba antiga: {aba}")

    # Restaura todas as abas locais no arquivo temp (preserva fórmulas e dados manuais)
    from openpyxl.cell.cell import MergedCell
    for nome_aba, dados in abas_locais.items():
        if nome_aba in wb_tmp.sheetnames:
            ws_t = wb_tmp[nome_aba]
            for i, linha in enumerate(dados, 1):
                for j, value in enumerate(linha, 1):
                    cell = ws_t.cell(row=i, column=j)
                    if not isinstance(cell, MergedCell):
                        cell.value = value
            print(f"   Restaurada: '{nome_aba}' ({len(dados)} linhas)")

    wb_tmp.save(arquivo_temp)

    # Cria apenas a aba do mês espelho
    kpis_esp_tuple = todos_kpis.get((ano_esp, mes_esp))
    if kpis_esp_tuple:
        kpis_df_esp, mes_nome_esp, _ = kpis_esp_tuple
        criar_nova_aba(arquivo_temp, kpis_df_esp, mes_nome_esp, ano_esp)
        # Espelha D e E na Planilha1 (nunca toca em outras colunas ou formulas)
        espelhar_planilha1(arquivo_temp, kpis_df_esp, mes_nome_esp, ano_esp)
    else:
        print(f"[AVISO] Sem dados para {mes_nome_esp}/{ano_esp} — aba nao criada.")

    # Passo 4: Gerar arquivo RH do mês anterior (fechado)
    kpis_mes_ref = todos_kpis.get((ano_ref, mes_ref))

    if kpis_mes_ref:
        kpis_df, mes_nome_ref, _ = kpis_mes_ref
        arquivo_rh = gerar_arquivo_rh(arquivo_temp, kpis_df, mes_nome_ref, ano_ref)
    else:
        print(f"[AVISO] Sem dados para {mes_nome_ref}/{ano_ref} — arquivo RH nao gerado.")
        arquivo_rh = None


    # Passo 5: Salvar no OneDrive
    if arquivo_rh:
        salvar_onedrive(arquivo_temp, arquivo_rh, mes_nome_ref, ano_ref)
    else:
        import shutil
        onedrive = detectar_onedrive()
        subpasta = os.path.join(
            onedrive if onedrive else os.path.expanduser("~") + "\\Desktop",
            "Controle KPIs Motoristas"
        )
        os.makedirs(subpasta, exist_ok=True)

        shutil.copy2(
            arquivo_temp,
            os.path.join(subpasta, f"Controle BRDRIVE - KPIs {date.today().year}.xlsx")
        )

    print()
    print("=" * 55)
    print("[OK] Concluido com sucesso!")
    print("=" * 55)


if __name__ == "__main__":
    main()
